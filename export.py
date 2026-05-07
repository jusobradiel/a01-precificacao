import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
import os
from datetime import datetime

BRAND = "0F3A4A"
ACCENT = "5DCAA5"
LIGHT = "E8F5F1"
GRAY = "F3F4F6"
WHITE = "FFFFFF"

def _fill(hex_color):
    return PatternFill("solid", fgColor=hex_color)

def _font(bold=False, color="1a1a1a", size=11):
    return Font(bold=bold, color=color, size=size, name="Calibri")

def _border():
    side = Side(style="thin", color="E5E7EB")
    return Border(left=side, right=side, top=side, bottom=side)

def _cell(ws, row, col, value, bg=WHITE, bold=False, color="1a1a1a", size=11, align="left", fmt=None):
    c = ws.cell(row=row, column=col, value=value)
    c.font = _font(bold=bold, color=color, size=size)
    c.fill = _fill(bg)
    c.alignment = Alignment(horizontal=align, vertical="center", wrap_text=True)
    c.border = _border()
    if fmt:
        c.number_format = fmt
    return c

def _header(ws, row, cols, bg=BRAND):
    for i, val in enumerate(cols, 1):
        _cell(ws, row, i, val, bg=bg, bold=True, color=WHITE, align="center")
    ws.row_dimensions[row].height = 24

def _money(val):
    return round(float(val or 0), 2)

def gerar_excel(cliente_id, cliente_nome, get_db_fn, get_hora_fn, get_custo_fn):
    hora = get_hora_fn(cliente_id)

    with get_db_fn() as conn:
        cfg = conn.execute('SELECT * FROM configuracoes WHERE cliente_id=?', (cliente_id,)).fetchone()
        fixos = conn.execute('SELECT * FROM custos_fixos WHERE cliente_id=? ORDER BY id', (cliente_id,)).fetchall()
        variaveis = conn.execute('SELECT * FROM custos_variaveis WHERE cliente_id=? ORDER BY id', (cliente_id,)).fetchall()
        insumos = conn.execute("SELECT * FROM insumos WHERE cliente_id=? AND nome != '__hora_clinica__' ORDER BY nome", (cliente_id,)).fetchall()
        protocolos = conn.execute('SELECT * FROM protocolos WHERE cliente_id=? ORDER BY nome', (cliente_id,)).fetchall()
        proto_itens = {}
        for p in protocolos:
            itens = conn.execute('''
                SELECT i.nome, i.custo_unitario, i.unidade, pi.quantidade_usada
                FROM protocolo_insumos pi JOIN insumos i ON pi.insumo_id=i.id
                WHERE pi.protocolo_id=?
            ''', (p['id'],)).fetchall()
            proto_itens[p['id']] = itens

    wb = openpyxl.Workbook()

    # ── ABA 1: CONFIGURAÇÕES ──────────────────────────────────────────────────
    ws = wb.active
    ws.title = "Configurações"
    ws.sheet_view.showGridLines = False
    ws.column_dimensions['A'].width = 32
    ws.column_dimensions['B'].width = 22

    # Título
    ws.merge_cells('A1:B1')
    _cell(ws, 1, 1, f"A'01 Precificação — {cliente_nome}", bg=BRAND, bold=True, color=WHITE, size=13, align="center")
    ws.row_dimensions[1].height = 30

    ws.merge_cells('A2:B2')
    _cell(ws, 2, 1, f"Gerado em {datetime.now().strftime('%d/%m/%Y às %H:%M')}", bg=GRAY, color="6b7280", size=10, align="center")
    ws.row_dimensions[2].height = 18

    # Hora clínica
    ws.row_dimensions[4].height = 26
    _cell(ws, 4, 1, "Hora Clínica Mínima", bg=ACCENT, bold=True, color=BRAND, align="center")
    _cell(ws, 4, 2, hora, bg=ACCENT, bold=True, color=BRAND, align="center", fmt='R$ #,##0.00')

    # Custos fixos
    _header(ws, 6, ["Custo Fixo", "Valor (R$)"])
    r = 7
    total_fixo = 0
    for f in fixos:
        bg = WHITE if r % 2 != 0 else GRAY
        _cell(ws, r, 1, f['descricao'], bg=bg)
        _cell(ws, r, 2, _money(f['valor']), bg=bg, fmt='R$ #,##0.00', align="right")
        total_fixo += _money(f['valor'])
        r += 1
    _cell(ws, r, 1, "TOTAL FIXO", bg=LIGHT, bold=True)
    _cell(ws, r, 2, total_fixo, bg=LIGHT, bold=True, fmt='R$ #,##0.00', align="right")
    r += 2

    _header(ws, r, ["Custo Variável", "Valor (R$)"])
    r += 1
    total_var = 0
    for v in variaveis:
        bg = WHITE if r % 2 != 0 else GRAY
        _cell(ws, r, 1, v['descricao'], bg=bg)
        _cell(ws, r, 2, _money(v['valor']), bg=bg, fmt='R$ #,##0.00', align="right")
        total_var += _money(v['valor'])
        r += 1
    _cell(ws, r, 1, "TOTAL VARIÁVEL", bg=LIGHT, bold=True)
    _cell(ws, r, 2, total_var, bg=LIGHT, bold=True, fmt='R$ #,##0.00', align="right")
    r += 2

    if cfg:
        _header(ws, r, ["Remuneração e Metas", "Valor"])
        r += 1
        items = [
            ("Pró-labore mensal", f"R$ {_money(cfg['pro_labore']):.2f}"),
            ("Lucro desejado", f"{_money(cfg['lucro_desejado']):.1f}%"),
            ("Horas clínicas por mês", int(cfg['horas_mes'])),
        ]
        for i, (k, v) in enumerate(items):
            bg = WHITE if i % 2 == 0 else GRAY
            _cell(ws, r, 1, k, bg=bg)
            _cell(ws, r, 2, v, bg=bg)
            r += 1

    # ── ABA 2: INSUMOS ────────────────────────────────────────────────────────
    ws2 = wb.create_sheet("Insumos")
    ws2.sheet_view.showGridLines = False
    for col, w in zip('ABCDE', [32, 15, 14, 18, 18]):
        ws2.column_dimensions[col].width = w

    ws2.merge_cells('A1:E1')
    _cell(ws2, 1, 1, f"Insumos — {cliente_nome}", bg=BRAND, bold=True, color=WHITE, size=13, align="center")
    ws2.row_dimensions[1].height = 28

    _header(ws2, 3, ["Matéria-prima", "Medida", "Quantidade", "Custo Unit. (R$)", "Custo Total (R$)"])

    for i, ins in enumerate(insumos):
        r = i + 4
        bg = WHITE if i % 2 == 0 else GRAY
        custo_total = _money(ins['custo_unitario']) * _money(ins['quantidade'])
        _cell(ws2, r, 1, ins['nome'], bg=bg)
        _cell(ws2, r, 2, ins['unidade'] or '', bg=bg, align="center")
        _cell(ws2, r, 3, int(ins['quantidade']), bg=bg, align="center")
        _cell(ws2, r, 4, _money(ins['custo_unitario']), bg=bg, fmt='R$ #,##0.00', align="right")
        _cell(ws2, r, 5, custo_total, bg=bg, fmt='R$ #,##0.00', align="right")

    # Hora clínica no final
    r = len(insumos) + 4
    _cell(ws2, r, 1, "Hora Clínica (automático)", bg=LIGHT, bold=True)
    _cell(ws2, r, 2, "unidade", bg=LIGHT, align="center")
    _cell(ws2, r, 3, 1, bg=LIGHT, align="center")
    _cell(ws2, r, 4, hora, bg=LIGHT, bold=True, fmt='R$ #,##0.00', align="right")
    _cell(ws2, r, 5, hora, bg=LIGHT, bold=True, fmt='R$ #,##0.00', align="right")

    # ── ABA 3: PROTOCOLOS ────────────────────────────────────────────────────
    ws3 = wb.create_sheet("Protocolos")
    ws3.sheet_view.showGridLines = False
    for col, w in zip('ABCDEFGHI', [28, 28, 12, 16, 16, 20, 20, 20, 12]):
        ws3.column_dimensions[col].width = w

    ws3.merge_cells('A1:I1')
    _cell(ws3, 1, 1, f"Protocolos — {cliente_nome}", bg=BRAND, bold=True, color=WHITE, size=13, align="center")
    ws3.row_dimensions[1].height = 28

    _header(ws3, 3, [
        "Protocolo", "Insumo usado", "Qtd usada", "Custo do item (R$)",
        "Custo total (R$)", "Cenário 1", "Cenário 2", "Cenário 3", "Cartão/Imp %"
    ])
    ws3.row_dimensions[3].height = 36

    r = 4
    for p in protocolos:
        custo_total = get_custo_fn(p['id'], hora)
        itens = proto_itens[p['id']]

        c1 = f"{p['cenario1_nome']}\nVenda: R${_money(p['cenario1_preco']):.2f}\nLucro: R${(_money(p['cenario1_preco'])-custo_total):.2f}\nMargem: {((_money(p['cenario1_preco'])-custo_total)/_money(p['cenario1_preco'])*100 if _money(p['cenario1_preco'])>0 else 0):.1f}%"
        c2 = f"{p['cenario2_nome']}\nVenda: R${_money(p['cenario2_preco']):.2f}\nLucro: R${(_money(p['cenario2_preco'])-custo_total):.2f}\nMargem: {((_money(p['cenario2_preco'])-custo_total)/_money(p['cenario2_preco'])*100 if _money(p['cenario2_preco'])>0 else 0):.1f}%"
        c3 = f"{p['cenario3_nome']}\nVenda: R${_money(p['cenario3_preco']):.2f}\nLucro: R${(_money(p['cenario3_preco'])-custo_total):.2f}\nMargem: {((_money(p['cenario3_preco'])-custo_total)/_money(p['cenario3_preco'])*100 if _money(p['cenario3_preco'])>0 else 0):.1f}%"
        cartao = f"{_money(p['cartao_imposto_pct']):.1f}%"

        rows_to_fill = itens if itens else [None]
        for idx, item in enumerate(rows_to_fill):
            is_first = idx == 0
            bg = LIGHT if is_first else (WHITE if idx % 2 == 0 else GRAY)

            _cell(ws3, r, 1, p['nome'] if is_first else "", bg=bg, bold=is_first)
            if item:
                custo_item = _money(item['custo_unitario']) * _money(item['quantidade_usada'])
                _cell(ws3, r, 2, item['nome'], bg=bg)
                _cell(ws3, r, 3, item['quantidade_usada'], bg=bg, align="center")
                _cell(ws3, r, 4, custo_item, bg=bg, fmt='R$ #,##0.00', align="right")
            else:
                _cell(ws3, r, 2, "(sem insumos)", bg=bg, color="9ca3af")
                _cell(ws3, r, 3, "", bg=bg)
                _cell(ws3, r, 4, "", bg=bg)

            _cell(ws3, r, 5, custo_total if is_first else "", bg=bg, bold=is_first, fmt='R$ #,##0.00' if is_first else None, align="right")
            _cell(ws3, r, 6, c1 if is_first else "", bg=bg)
            _cell(ws3, r, 7, c2 if is_first else "", bg=bg)
            _cell(ws3, r, 8, c3 if is_first else "", bg=bg)
            _cell(ws3, r, 9, cartao if is_first else "", bg=bg, align="center")

            if is_first:
                ws3.row_dimensions[r].height = 60
            r += 1

        # linha separadora
        for col in range(1, 10):
            ws3.cell(r, col).fill = _fill("E5E7EB")
        r += 1

    return wb
